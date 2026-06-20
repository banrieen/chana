#!/usr/bin/env python3
"""
Chana Requirement Manager — 需求管理CLI工具
================================================
功能: 数据验证 / 导入导出 / 查询筛选 / 统计报表 / 变更追踪

Usage:
  python3 req_manager.py validate                    # 验证CSV数据完整性
  python3 req_manager.py stats                       # 生成统计概览
  python3 req_manager.py query --status 规划中        # 按条件查询
  python3 req_manager.py query --milestone MVP       # 按里程碑筛选
  python3 req_manager.py query --kano M --moscow Must # 多维筛选
  python3 req_manager.py update M01 --status 研发中   # 更新需求状态
  python3 req_manager.py update M01 --owner Alice     # 分配负责人
  python3 req_manager.py export --format md          # 导出Markdown
  python3 req_manager.py export --format json        # 导出JSON
  python3 req_manager.py export --format xlsx        # 导出Excel
  python3 req_manager.py import --file new_reqs.csv  # 导入CSV
  python3 req_manager.py log                         # 查看变更历史
  python3 req_manager.py report                      # 生成HTML报表
  python3 req_manager.py deps M01                    # 查看依赖链
"""

import csv
import json
import argparse
import os
import sys
import hashlib
from datetime import datetime, date
from pathlib import Path

# ── Paths ──────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
SCHEMA_FILE = BASE_DIR / "schema.json"

# Data dir for all read/write operations
DATA_DIR = Path("/Users/bianshengzhi/WorkBuddy/outputs/chana")
DATA_DIR.mkdir(parents=True, exist_ok=True)

# If CSV not in data dir, copy from project
CSV_FILE = DATA_DIR / "requirements.csv"
CHANGELOG_FILE = DATA_DIR / "changelog.csv"
if not CSV_FILE.exists() and (BASE_DIR / "requirements.csv").exists():
    import shutil
    shutil.copy(BASE_DIR / "requirements.csv", CSV_FILE)

# ── Schema ─────────────────────────────────────────────
def load_schema():
    with open(SCHEMA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def load_requirements():
    if not CSV_FILE.exists():
        print(f"❌ CSV文件不存在: {CSV_FILE}")
        sys.exit(1)
    with open(CSV_FILE, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader)

def save_requirements(reqs):
    if not reqs:
        print("⚠️ 无数据可保存")
        return
    fieldnames = [k for k in reqs[0].keys() if k is not None]
    with open(CSV_FILE, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(reqs)

# ── Validation ─────────────────────────────────────────
def calc_priority(kano: str, moscow: str) -> int:
    schema = load_schema()
    key = f"{kano}+{moscow}"
    return schema["priority_matrix"].get(key, 5)

def validate_requirements(reqs=None):
    schema = load_schema()
    if reqs is None:
        reqs = load_requirements()

    errors = []
    warnings = []
    seen_ids = set()

    # Build valid enum sets
    enum_fields = {}
    for field in schema["fields"]:
        if field["type"] == "enum":
            enum_fields[field["name"]] = set(field["values"])

    for i, req in enumerate(reqs):
        row_num = i + 2  # +1 header, +1 0-indexed
        req_id = req.get("req_id", "?")

        # Check required fields
        for field in schema["fields"]:
            fname = field["name"]
            if field.get("required") and not req.get(fname, "").strip():
                errors.append(f"行{row_num} [{req_id}]: 必填字段 '{fname}' 为空")

        # Check unique ID
        if req_id in seen_ids:
            errors.append(f"行{row_num}: 重复需求ID '{req_id}'")
        seen_ids.add(req_id)

        # Check ID pattern
        import re
        if "pattern" in {"req_id": schema["fields"][0]}.get("req_id", {}):
            pass  # simplified

        # Check enum values
        for fname, valid_values in enum_fields.items():
            val = req.get(fname, "").strip()
            if val and val not in valid_values:
                errors.append(
                    f"行{row_num} [{req_id}]: '{fname}' 值 '{val}' 无效，"
                    f"可选: {', '.join(sorted(valid_values))}"
                )

        # Check priority consistency
        kano = req.get("kano", "")
        moscow = req.get("moscow", "")
        if kano and moscow:
            expected_p = calc_priority(kano, moscow)
            actual_p = req.get("priority", "0")
            try:
                if int(actual_p) != expected_p:
                    warnings.append(
                        f"行{row_num} [{req_id}]: priority={actual_p} "
                        f"应为 {expected_p} (kano={kano}, moscow={moscow})"
                    )
            except ValueError:
                errors.append(f"行{row_num} [{req_id}]: priority 不是有效数字")

        # Check effort_est range
        try:
            effort = float(req.get("effort_est", "0"))
            if effort < 0.5 or effort > 30:
                warnings.append(f"行{row_num} [{req_id}]: effort_est={effort} 超出建议范围 0.5~30")
        except ValueError:
            errors.append(f"行{row_num} [{req_id}]: effort_est 不是有效数字")

        # Check status transition validity
        status = req.get("status", "")
        if status and status not in schema["status_transitions"]:
            errors.append(f"行{row_num} [{req_id}]: 未知状态 '{status}'")

        # Check dependencies exist
        deps = req.get("dependencies", "").strip()
        if deps:
            dep_list = [d.strip() for d in deps.split(";") if d.strip()]
            for dep in dep_list:
                if dep not in [r.get("req_id") for r in reqs]:
                    warnings.append(f"行{row_num} [{req_id}]: 依赖需求 '{dep}' 不存在")

    # Summary
    print("=" * 60)
    print(f"  验证结果: {len(reqs)} 条需求")
    print(f"  ❌ 错误: {len(errors)}")
    print(f"  ⚠️ 警告: {len(warnings)}")
    print("=" * 60)

    if errors:
        print("\n❌ 错误:")
        for e in errors:
            print(f"  • {e}")
    if warnings:
        print("\n⚠️ 警告:")
        for w in warnings:
            print(f"  • {w}")

    if not errors and not warnings:
        print("\n✅ 全部通过验证")

    return len(errors) == 0

# ── Statistics ──────────────────────────────────────────
def show_stats():
    reqs = load_requirements()
    total = len(reqs)

    by_status = {}
    by_milestone = {}
    by_kano = {}
    by_moscow = {}
    by_owner = {}
    total_effort = 0.0

    for r in reqs:
        s = r.get("status", "未知")
        by_status[s] = by_status.get(s, 0) + 1

        m = r.get("milestone", "未知")
        by_milestone[m] = by_milestone.get(m, 0) + 1

        k = r.get("kano", "?")
        by_kano[k] = by_kano.get(k, 0) + 1

        mo = r.get("moscow", "?")
        by_moscow[mo] = by_moscow.get(mo, 0) + 1

        owner = r.get("owner", "") or "未分配"
        by_owner[owner] = by_owner.get(owner, 0) + 1

        try:
            total_effort += float(r.get("effort_est", 0))
        except ValueError:
            pass

    done = by_status.get("已交付", 0)
    in_progress = by_status.get("研发中", 0) + by_status.get("测试中", 0)
    progress_pct = (done / total * 100) if total else 0

    print()
    print("╔" + "═" * 58 + "╗")
    print("║" + " 刹那 (Chana) — 需求管理统计概览".center(46) + "║")
    print("╠" + "═" * 58 + "╣")

    print(f"║  总需求数: {total:<42} ║")
    print(f"║  总工作量: {total_effort:.1f} 人天{'':>34} ║")
    print(f"║  完成进度: {done}/{total} ({progress_pct:.1f}%){'':>24} ║")
    print(f"║  进行中:   {in_progress:<42} ║")
    print("╠" + "═" * 58 + "╣")

    print("║  按状态:                                              ║")
    for s in ["规划中", "研发中", "测试中", "已交付", "已搁置"]:
        cnt = by_status.get(s, 0)
        bar = "█" * cnt + "░" * (max(by_status.values()) - cnt if by_status else 0)
        print(f"║    {s:6s}: {cnt:>2}  {bar:<28} ║")

    print("╠" + "═" * 58 + "╣")
    print("║  按里程碑:                                            ║")
    for m in ["Demo", "MVP", "1.0", "1.5", "2.0"]:
        cnt = by_milestone.get(m, 0)
        ms_effort = sum(
            float(r.get("effort_est", 0))
            for r in reqs
            if r.get("milestone") == m
        )
        try:
            ms_done = sum(1 for r in reqs if r.get("milestone") == m and r.get("status") == "已交付")
        except:
            ms_done = 0
        print(f"║    {m:5s}: {cnt:>2} 条  {ms_effort:>5.1f}人天  完成 {ms_done}/{cnt:<10} ║")

    print("╠" + "═" * 58 + "╣")
    print("║  按KANO:    " + "  ".join(
        f"{k}={by_kano.get(k,0)}" for k in ["M","O","A"]
    ).ljust(44) + "║")
    print("║  按MoSCoW:  " + "  ".join(
        f"{m}={by_moscow.get(m,0)}" for m in ["Must","Should","Could"]
    ).ljust(44) + "║")

    print("╠" + "═" * 58 + "╣")
    print("║  按负责人:                                            ║")
    for owner, cnt in sorted(by_owner.items(), key=lambda x: -x[1]):
        print(f"║    {owner:12s}: {cnt:<36} ║")

    print("╚" + "═" * 58 + "╝")

# ── Query ───────────────────────────────────────────────
def query_reqs(args):
    reqs = load_requirements()
    results = reqs

    if args.status:
        results = [r for r in results if r.get("status") == args.status]
    if args.milestone:
        results = [r for r in results if r.get("milestone") == args.milestone]
    if args.kano:
        results = [r for r in results if r.get("kano") == args.kano]
    if args.moscow:
        results = [r for r in results if r.get("moscow") == args.moscow]
    if args.owner:
        results = [r for r in results if args.owner.lower() in r.get("owner", "").lower()]
    if args.tag:
        results = [r for r in results if args.tag.lower() in r.get("tags", "").lower()]
    if args.id:
        results = [r for r in results if args.id.upper() in r.get("req_id", "").upper()]

    if args.sort:
        sort_field = args.sort
        results.sort(key=lambda r: r.get(sort_field, ""))

    print(f"\n查询结果: {len(results)} / {len(reqs)} 条\n")
    print(f"{'ID':6s} {'名称':16s} {'KANO':5s} {'MoSCoW':8s} {'MS':5s} {'状态':6s} {'P':2s} {'工作量':6s} {'负责人':10s}")
    print("-" * 80)
    for r in results:
        print(
            f"{r.get('req_id',''):6s} "
            f"{r.get('name','')[:14]:16s} "
            f"{r.get('kano',''):5s} "
            f"{r.get('moscow',''):8s} "
            f"{r.get('milestone',''):5s} "
            f"{r.get('status',''):6s} "
            f"{r.get('priority',''):2s} "
            f"{r.get('effort_est',''):>5s}d "
            f"{r.get('owner','') or '-':10s}"
        )

    if args.verbose:
        print("\n" + "=" * 80)
        for r in results:
            print(f"\n📌 {r['req_id']} — {r['name']}")
            print(f"   描述: {r.get('description','')}")
            print(f"   验收: {r.get('acceptance_criteria','')}")
            print(f"   依赖: {r.get('dependencies','无')}")
            print(f"   风险: {r.get('risk_ids','无')}")
            print(f"   标签: {r.get('tags','无')}")
            print(f"   版本: {r.get('version','')} | 更新: {r.get('updated_at','')}")

# ── Update ──────────────────────────────────────────────
def update_req(args):
    reqs = load_requirements()
    found = False

    for r in reqs:
        if r["req_id"].upper() == args.id.upper():
            found = True
            old_values = {}

            if args.status:
                schema = load_schema()
                old_status = r.get("status", "")
                transitions = schema["status_transitions"]
                if old_status in transitions:
                    valid_next = transitions[old_status]
                    if args.status not in valid_next and old_status != args.status:
                        print(f"❌ 状态流转无效: {old_status} → {args.status}")
                        print(f"   允许: {', '.join(valid_next) if valid_next else '(终态)'}")
                        return
                old_values["status"] = old_status
                r["status"] = args.status

            if args.owner:
                old_values["owner"] = r.get("owner", "")
                r["owner"] = args.owner

            if args.milestone:
                old_values["milestone"] = r.get("milestone", "")
                r["milestone"] = args.milestone

            if args.moscow:
                old_values["moscow"] = r.get("moscow", "")
                r["moscow"] = args.moscow
                # Recalculate priority
                r["priority"] = str(calc_priority(r.get("kano",""), args.moscow))

            if args.effort:
                old_values["effort_est"] = r.get("effort_est", "")
                r["effort_est"] = str(args.effort)

            if args.notes:
                old_values["notes"] = r.get("notes", "")
                r["notes"] = args.notes

            r["updated_at"] = datetime.now().strftime("%Y-%m-%d")

            # Record change
            log_change(args.id, old_values, {k: r.get(k, "") for k in old_values})

            print(f"✅ 已更新 {args.id}:")
            for k, old_v in old_values.items():
                print(f"   {k}: {old_v} → {r.get(k, '')}")
            break

    if not found:
        print(f"❌ 未找到需求: {args.id}")
        return

    save_requirements(reqs)
    print(f"\n💾 已保存到 {CSV_FILE}")

# ── Changelog ───────────────────────────────────────────
def log_change(req_id, old_values, new_values):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    changes = []
    for k in old_values:
        changes.append(f"{k}: {old_values[k]} → {new_values.get(k, '')}")
    change_desc = "; ".join(changes)

    file_exists = CHANGELOG_FILE.exists()
    with open(CHANGELOG_FILE, "a", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["timestamp", "req_id", "change", "operator"])
        writer.writerow([timestamp, req_id, change_desc, os.getenv("USER", "unknown")])

def show_log(args):
    if not CHANGELOG_FILE.exists():
        print("📝 暂无变更记录")
        return

    with open(CHANGELOG_FILE, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        logs = list(reader)

    if args.id:
        logs = [l for l in logs if l.get("req_id", "").upper() == args.id.upper()]

    print(f"\n变更历史: {len(logs)} 条记录\n")
    print(f"{'时间':21s} {'需求ID':6s} {'变更内容':50s} {'操作人':10s}")
    print("-" * 90)
    for l in logs:
        print(f"{l.get('timestamp',''):21s} {l.get('req_id',''):6s} {l.get('change','')[:48]:50s} {l.get('operator',''):10s}")

# ── Export ──────────────────────────────────────────────
def export_reqs(args):
    reqs = load_requirements()
    fmt = args.format

    if fmt == "json":
        out = DATA_DIR / "requirements.json"
        with open(out, "w", encoding="utf-8") as f:
            json.dump(reqs, f, ensure_ascii=False, indent=2)
        print(f"✅ 导出 JSON: {out} ({len(reqs)} 条)")

    elif fmt == "md":
        out = DATA_DIR / "requirements.md"
        with open(out, "w", encoding="utf-8") as f:
            f.write("# 刹那 (Chana) — 需求管理宽表\n\n")
            f.write(f"> 导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M')} | 共 {len(reqs)} 条需求\n\n")
            f.write("| ID | 名称 | 描述 | KANO | MoSCoW | 里程碑 | 状态 | 优先级 | 负责人 | 工作量 | 验收标准 | 依赖 | 标签 |\n")
            f.write("|----|------|------|------|--------|--------|------|--------|--------|--------|----------|------|------|\n")
            for r in reqs:
                f.write(f"| {r['req_id']} | {r['name']} | {r.get('description','')} | {r.get('kano','')} | {r.get('moscow','')} | {r.get('milestone','')} | {r.get('status','')} | {r.get('priority','')} | {r.get('owner','')} | {r.get('effort_est','')}d | {r.get('acceptance_criteria','')} | {r.get('dependencies','')} | {r.get('tags','')} |\n")
        print(f"✅ 导出 Markdown: {out} ({len(reqs)} 条)")

    elif fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("❌ 需要 openpyxl 库: pip install openpyxl")
            return

        out = DATA_DIR / "requirements.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "需求宽表"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Status colors
        status_colors = {
            "规划中": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
            "研发中": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
            "测试中": PatternFill(start_color="FFCC80", end_color="FFCC80", fill_type="solid"),
            "已交付": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
            "已搁置": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        }

        headers = list(reqs[0].keys()) if reqs else []
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, req in enumerate(reqs, 2):
            status = req.get("status", "")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col, value=req.get(h, ""))
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if h == "status" and status in status_colors:
                    cell.fill = status_colors[status]

        # Column widths
        col_widths = {
            "req_id": 8, "name": 18, "description": 35, "kano": 6, "moscow": 9,
            "milestone": 9, "status": 8, "priority": 6, "owner": 10, "effort_est": 8,
            "acceptance_criteria": 35, "dependencies": 12, "risk_ids": 10, "tags": 15,
            "created_at": 12, "updated_at": 12, "version": 8, "notes": 20,
        }
        for col, h in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = col_widths.get(h, 12)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        wb.save(out)
        print(f"✅ 导出 Excel: {out} ({len(reqs)} 条)")

    elif fmt == "csv":
        print(f"ℹ️ CSV 已是主格式: {CSV_FILE}")

    else:
        print(f"❌ 不支持的格式: {fmt} (可选: json/md/xlsx/csv)")

# ── Import ──────────────────────────────────────────────
def import_reqs(args):
    filepath = Path(args.file)
    if not filepath.exists():
        print(f"❌ 文件不存在: {filepath}")
        return

    existing = load_requirements()
    existing_ids = {r["req_id"] for r in existing}

    new_reqs = []
    if filepath.suffix == ".csv":
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            new_reqs = list(reader)
    elif filepath.suffix in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                new_reqs.append(dict(zip(headers, row)))
        except ImportError:
            print("❌ 需要 openpyxl: pip install openpyxl")
            return
    else:
        print(f"❌ 不支持的文件格式: {filepath.suffix}")
        return

    added = 0
    skipped = 0
    for req in new_reqs:
        rid = req.get("req_id", "").strip()
        if not rid:
            skipped += 1
            continue
        if rid in existing_ids:
            print(f"  ⏭️  跳过重复: {rid}")
            skipped += 1
            continue
        # Ensure required fields have defaults
        req.setdefault("status", "规划中")
        req.setdefault("priority", str(calc_priority(
            req.get("kano", "M"), req.get("moscow", "Must")
        )))
        req.setdefault("created_at", datetime.now().strftime("%Y-%m-%d"))
        req.setdefault("updated_at", datetime.now().strftime("%Y-%m-%d"))
        req.setdefault("version", "v1.0")
        existing.append(req)
        added += 1
        print(f"  ✅ 新增: {rid} — {req.get('name', '')}")

    save_requirements(existing)
    print(f"\n📊 导入完成: 新增 {added} 条, 跳过 {skipped} 条, 总计 {len(existing)} 条")

# ── Dependencies ────────────────────────────────────────
def show_deps(args):
    reqs = load_requirements()
    req_map = {r["req_id"]: r for r in reqs}

    target = args.id.upper()
    if target not in req_map:
        print(f"❌ 未找到需求: {target}")
        return

    print(f"\n📌 {target} — {req_map[target]['name']}")
    print(f"   {req_map[target].get('description','')}")

    # Direct dependencies
    deps = req_map[target].get("dependencies", "").strip()
    if deps:
        print(f"\n📋 直接依赖:")
        for dep in deps.split(";"):
            dep = dep.strip()
            if dep in req_map:
                d = req_map[dep]
                print(f"   → {dep}: {d['name']} [{d.get('status','')}]")
            else:
                print(f"   → {dep}: (未找到)")

    # Who depends on this
    dependents = []
    for r in reqs:
        r_deps = r.get("dependencies", "")
        if target in [d.strip() for d in r_deps.split(";") if d.strip()]:
            dependents.append(r)

    if dependents:
        print(f"\n📋 被以下需求依赖:")
        for r in dependents:
            print(f"   ← {r['req_id']}: {r['name']} [{r.get('status','')}]")

    # Transitive deps (depth 3)
    print(f"\n📋 传递依赖链 (max depth=3):")
    visited = set()
    def trace(rid, depth=0):
        if rid in visited or depth > 3:
            return
        visited.add(rid)
        if rid in req_map:
            r_deps = req_map[rid].get("dependencies", "").strip()
            if r_deps:
                for dep in r_deps.split(";"):
                    dep = dep.strip()
                    prefix = "  " * (depth + 1)
                    status_icon = "✅" if req_map.get(dep, {}).get("status") == "已交付" else "⏳"
                    print(f"   {prefix}{rid} → {dep}: {req_map.get(dep, {}).get('name', '?')} {status_icon}")
                    trace(dep, depth + 1)
    trace(target)

# ── Report ──────────────────────────────────────────────
def generate_report():
    reqs = load_requirements()
    report_file = DATA_DIR / "07-需求管理报表.html"

    by_milestone = {}
    by_status = {}
    by_kano = {}
    total_effort = 0
    done_effort = 0

    for r in reqs:
        ms = r.get("milestone", "未知")
        by_milestone.setdefault(ms, {"total": 0, "done": 0, "effort": 0})
        by_milestone[ms]["total"] += 1
        by_milestone[ms]["effort"] += float(r.get("effort_est", 0))
        if r.get("status") == "已交付":
            by_milestone[ms]["done"] += 1
            done_effort += float(r.get("effort_est", 0))

        s = r.get("status", "未知")
        by_status[s] = by_status.get(s, 0) + 1

        k = r.get("kano", "?")
        by_kano[k] = by_kano.get(k, 0) + 1

        total_effort += float(r.get("effort_est", 0))

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>刹那 — 需求统计报表</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<style>
  body {{ font-family: -apple-system, sans-serif; background: #f5f5f5; margin: 0; padding: 20px; }}
  .container {{ max-width: 900px; margin: 0 auto; }}
  h1 {{ color: #1a1a2e; text-align: center; }}
  .summary {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin: 20px 0; }}
  .card {{ background: white; border-radius: 12px; padding: 20px; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
  .card .num {{ font-size: 2em; font-weight: bold; color: #2B579A; }}
  .card .label {{ color: #666; margin-top: 4px; }}
  .chart-row {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin: 20px 0; }}
  .chart-box {{ background: white; border-radius: 12px; padding: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
  table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
  th {{ background: #2B579A; color: white; padding: 10px; text-align: left; }}
  td {{ padding: 8px 10px; border-bottom: 1px solid #eee; }}
  .status-规划中 {{ color: #1976D2; }}
  .status-研发中 {{ color: #F57F17; }}
  .status-测试中 {{ color: #E65100; }}
  .status-已交付 {{ color: #2E7D32; font-weight: bold; }}
  .progress-bar {{ background: #e0e0e0; border-radius: 4px; height: 20px; overflow: hidden; }}
  .progress-fill {{ height: 100%; background: #4CAF50; display: flex; align-items: center; justify-content: center; color: white; font-size: 12px; }}
</style>
</head>
<body>
<div class="container">
  <h1>⚡ 刹那 (Chana) — 需求统计报表</h1>
  <p style="text-align:center;color:#666;">生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>

  <div class="summary">
    <div class="card"><div class="num">{len(reqs)}</div><div class="label">总需求数</div></div>
    <div class="card"><div class="num">{by_status.get('已交付',0)}</div><div class="label">已交付</div></div>
    <div class="card"><div class="num">{total_effort:.1f}</div><div class="label">总工作量(人天)</div></div>
    <div class="card"><div class="num">{(by_status.get('已交付',0)/len(reqs)*100):.1f}%</div><div class="label">完成率</div></div>
  </div>

  <div class="chart-row">
    <div class="chart-box"><canvas id="statusChart"></canvas></div>
    <div class="chart-box"><canvas id="milestoneChart"></canvas></div>
  </div>
  <div class="chart-row">
    <div class="chart-box"><canvas id="kanoChart"></canvas></div>
    <div class="chart-box"><canvas id="effortChart"></canvas></div>
  </div>

  <h3>里程碑进度</h3>
  <table>
    <tr><th>里程碑</th><th>需求数</th><th>已完成</th><th>工作量</th><th>进度</th></tr>
"""
    for ms in ["Demo", "MVP", "1.0", "1.5", "2.0"]:
        d = by_milestone.get(ms, {"total": 0, "done": 0, "effort": 0})
        pct = (d["done"] / d["total"] * 100) if d["total"] else 0
        html += f"""    <tr>
      <td><b>{ms}</b></td><td>{d['total']}</td><td>{d['done']}</td><td>{d['effort']:.1f}d</td>
      <td><div class="progress-bar"><div class="progress-fill" style="width:{pct}%">{pct:.0f}%</div></div></td>
    </tr>
"""
    html += """  </table>

  <h3>需求明细</h3>
  <table>
    <tr><th>ID</th><th>名称</th><th>KANO</th><th>MoSCoW</th><th>里程碑</th><th>状态</th><th>工作量</th><th>负责人</th></tr>
"""
    for r in reqs:
        status = r.get("status", "")
        html += f"""    <tr>
      <td>{r['req_id']}</td><td>{r['name']}</td><td>{r.get('kano','')}</td>
      <td>{r.get('moscow','')}</td><td>{r.get('milestone','')}</td>
      <td class="status-{status}">{status}</td>
      <td>{r.get('effort_est','')}d</td><td>{r.get('owner','') or '-'}</td>
    </tr>
"""
    html += """  </table>
</div>

<script>
"""
    # Status chart
    statuses = ["规划中", "研发中", "测试中", "已交付", "已搁置"]
    status_data = [by_status.get(s, 0) for s in statuses]
    status_colors = ["#42A5F5", "#FFEE58", "#FFA726", "#66BB6A", "#EF5350"]

    import json

    # Build JS chart configs as Python dicts, then serialize
    chart_configs = [
        {
            "id": "statusChart",
            "type": "doughnut",
            "data": {
                "labels": statuses,
                "datasets": [{"data": status_data, "backgroundColor": status_colors}]
            },
            "options": {"plugins": {"title": {"display": True, "text": "需求状态分布"}}}
        },
        {
            "id": "milestoneChart",
            "type": "bar",
            "data": {
                "labels": list(by_milestone.keys()),
                "datasets": [
                    {"label": "需求数", "data": [by_milestone[m]["total"] for m in by_milestone], "backgroundColor": "#42A5F5"},
                    {"label": "已完成", "data": [by_milestone[m]["done"] for m in by_milestone], "backgroundColor": "#66BB6A"}
                ]
            },
            "options": {"plugins": {"title": {"display": True, "text": "里程碑进度"}}}
        },
        {
            "id": "kanoChart",
            "type": "pie",
            "data": {
                "labels": ["基本型(M)", "期望型(O)", "兴奋型(A)"],
                "datasets": [{"data": [by_kano.get(k, 0) for k in ["M", "O", "A"]], "backgroundColor": ["#EF5350", "#FFA726", "#66BB6A"]}]
            },
            "options": {"plugins": {"title": {"display": True, "text": "KANO分类"}}}
        },
        {
            "id": "effortChart",
            "type": "bar",
            "data": {
                "labels": list(by_milestone.keys()),
                "datasets": [{"label": "工作量(人天)", "data": [by_milestone[m]["effort"] for m in by_milestone], "backgroundColor": "#7E57C2"}]
            },
            "options": {"indexAxis": "y", "plugins": {"title": {"display": True, "text": "工作量分布"}}}
        }
    ]

    for cfg in chart_configs:
        chart_id = cfg.pop("id")
        js_cfg = json.dumps(cfg, ensure_ascii=False)
        # Convert Python True/False/None to JS true/false/null
        js_cfg = js_cfg.replace("true", "true").replace("false", "false").replace("null", "null")
        html += f"new Chart(document.getElementById('{chart_id}'), {js_cfg});\n"
    html += "</script>\n</body>\n</html>"

    with open(report_file, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"✅ 报表已生成: {report_file}")
    print(f"   用浏览器打开查看: file://{report_file}")

# ── Main ────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="刹那 (Chana) 需求管理CLI工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="示例:\n"
               "  python3 req_manager.py validate\n"
               "  python3 req_manager.py stats\n"
               "  python3 req_manager.py query --milestone MVP\n"
               "  python3 req_manager.py update M01 --status 研发中 --owner Alice\n"
               "  python3 req_manager.py export --format xlsx\n"
               "  python3 req_manager.py report\n"
    )
    subparsers = parser.add_subparsers(dest="command", help="可用命令")

    # validate
    subparsers.add_parser("validate", help="验证CSV数据完整性")

    # stats
    subparsers.add_parser("stats", help="生成统计概览")

    # query
    q_parser = subparsers.add_parser("query", help="查询筛选需求")
    q_parser.add_argument("--status", help="按状态筛选")
    q_parser.add_argument("--milestone", help="按里程碑筛选")
    q_parser.add_argument("--kano", help="按KANO分类筛选 (M/O/A)")
    q_parser.add_argument("--moscow", help="按MoSCoW筛选 (Must/Should/Could)")
    q_parser.add_argument("--owner", help="按负责人筛选")
    q_parser.add_argument("--tag", help="按标签筛选")
    q_parser.add_argument("--id", help="按需求ID搜索")
    q_parser.add_argument("--sort", help="排序字段")
    q_parser.add_argument("-v", "--verbose", action="store_true", help="显示详细信息")

    # update
    u_parser = subparsers.add_parser("update", help="更新需求")
    u_parser.add_argument("id", help="需求ID, 如 M01")
    u_parser.add_argument("--status", help="新状态")
    u_parser.add_argument("--owner", help="负责人")
    u_parser.add_argument("--milestone", help="里程碑")
    u_parser.add_argument("--moscow", help="MoSCoW优先级")
    u_parser.add_argument("--effort", type=float, help="工作量估算")
    u_parser.add_argument("--notes", help="备注")

    # export
    e_parser = subparsers.add_parser("export", help="导出需求")
    e_parser.add_argument("--format", choices=["csv", "json", "md", "xlsx"], default="json", help="导出格式")

    # import
    i_parser = subparsers.add_parser("import", help="导入需求")
    i_parser.add_argument("--file", required=True, help="导入文件路径 (CSV/XLSX)")

    # log
    l_parser = subparsers.add_parser("log", help="查看变更历史")
    l_parser.add_argument("--id", help="按需求ID筛选")

    # deps
    d_parser = subparsers.add_parser("deps", help="查看依赖关系")
    d_parser.add_argument("id", help="需求ID, 如 M01")

    # report
    subparsers.add_parser("report", help="生成HTML报表")

    args = parser.parse_args()

    if args.command == "validate":
        validate_requirements()
    elif args.command == "stats":
        show_stats()
    elif args.command == "query":
        query_reqs(args)
    elif args.command == "update":
        update_req(args)
    elif args.command == "export":
        export_reqs(args)
    elif args.command == "import":
        import_reqs(args)
    elif args.command == "log":
        show_log(args)
    elif args.command == "deps":
        show_deps(args)
    elif args.command == "report":
        generate_report()
    else:
        parser.print_help()

if __name__ == "__main__":
    main()
